from ultralytics import YOLO
import constants
from constants import write_to_pdf,normalized_to_pdf_rect_px_to_pt,annotate_doc
import easyocr
import pandas as pd
from os import listdir
from os.path import isfile, join
from pathlib import Path
from PIL import Image
import re
import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
import cv2
import numpy as np
import fitz
from pypdf import PdfReader, PdfWriter
from pypdf.annotations import FreeText




doc_reader = easyocr.Reader(['en'])

def detect_instrument(file_path):
    prediction_dict = dict()
    IMAGE_PATH = Path(file_path)
    image_path_list = list(IMAGE_PATH.iterdir())
    for q,file in enumerate(image_path_list):
        prediction_list = []
        img = cv2.imread(file, cv2.IMREAD_COLOR) 
        prediction_list.append(img)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) 
        gray_blurred = cv2.blur(gray, (7, 7)) 
        detected_circles = cv2.HoughCircles(gray_blurred,  
                   cv2.HOUGH_GRADIENT, 1.5, 40, param1 = 90 , 
               param2 = 40, minRadius = 45, maxRadius = 51) 
        print(f"Completed detections on {q+1} out of {len(image_path_list)} documents")

        if detected_circles is not None: 
            detected_circles = np.uint16(np.around(detected_circles)) 
            for pt in detected_circles[0, :]: 
                a, b, r = pt[0], pt[1], pt[2] 
                top_left = (a - r, b - r)
                bottom_right = (a + r, b + r)
                prediction_list.append([top_left,bottom_right])
            prediction_dict[file.name] = prediction_list
    print(len(prediction_dict))
    return prediction_dict

def get_bbox(prediction_dict):
    cropped_images_dict = dict()
    for pred in prediction_dict:
        cropped_image_list = []
        for bbox in prediction_dict[pred][1:]:
            x1,y1,x2,y2 = int(bbox[0][0]),int(bbox[0][1]),int(bbox[1][0]),int(bbox[1][1])
            image_tensor = prediction_dict[pred][0]
            cropped_image = image_tensor[y1:y2,x1:x2]

            image_height, image_width = image_tensor.shape[:2]

            bx = (x1/image_width) # this is in pixels
            by = (y1/image_height)
            bw = (x2/image_width)
            bh = (y2/image_height)
            bounding_box = (bx,by,bw,bh)

            # probably numpy uses a top left handed coordinate system
            # pypdf uses a top left handed coordinate system (probably) 

            # for pypdf x is the vertical, y is the horizontal

            cropped_image_list.append((cropped_image,bounding_box))
        cropped_images_dict[pred] = cropped_image_list
    return cropped_images_dict

def detect_text(imgs_dict):
    text_dict = dict()
    for w,image in enumerate(imgs_dict):
        result_list = []
        image_list = imgs_dict[image]
        for pic in image_list:
            result = doc_reader.readtext(pic[0],detail=0)
            result_string = ''
            for i in result:
                result_string += i
            result_list.append((result_string,pic[0],pic[1]))
        text_dict[image] = result_list

        print("\n")
        print(f"Completed detections on {w+1} out of {len(imgs_dict)} documents")
    return text_dict

def format_text(text_dict):
    formatted_text = dict()
    for image in text_dict:
        result_list = []
        text_list = text_dict[image]
        for texts in text_list:
            new_text = re.sub(r'[^A-Za-z0-9]', '', texts[0])
            pattern = r'^\d{2,3}[A-Z]{2,4}[A-Z0-9]{3,4}$'
            matches = re.findall(pattern, new_text)
            if matches:
                format_text = re.sub(r'([A-Z]+)(\d.*)$', r'\1-\2', new_text)
                result_list.append((format_text,texts[2]))
            else:
                result_list.append(('',texts[2]))
            
        formatted_text[image] = result_list
    return formatted_text

def search_sheet(formatted_text_dict,file_path):

    column_names = ["temperature","pressure or differential pressure","flowrate","alarm"]
    search_text_columns = ["normal operating","normal","low low",'high high','high','low',"uom"]
    
    wb = openpyxl.load_workbook(file_path)
    search_queries = []
    result_row = []
    result_column = []
    num = 0

    for image in formatted_text_dict:
        for i,texts in enumerate(formatted_text_dict[image]):
            text = texts[0]
            if text != '':
                search_queries.append({
                    "image_source":image,
                    "positional_identifier":i,
                    "search_query":text
                })
    

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        header_row = next(sheet.iter_rows(min_row=8, max_row=8))  
        target_col_idx = None
        hold = {}
        hold_list = []

        for column_name in column_names:
            for cell in header_row:
                if cell.value and column_name.lower().replace(" ","") == str(cell.value).lower().replace("\n","").replace(" ",""):
                    target_col_idx = [cell.column,column_name] 
                    hold_list.append(target_col_idx)
                    hold[f"{sheet_name}"] = hold_list
                    break

        if target_col_idx:
            for column in sheet.iter_cols(min_row=8, min_col=1, max_col=target_col_idx[0]+4):
                for cell in column:
                    for search_text_column in search_text_columns:
                        if cell.value and search_text_column.replace(" ","") == str(cell.value).replace("\n","").replace(" ","").lower():
                            exact_c = ''.join(re.findall(r'[A-Za-z]',cell.coordinate))
                            exact_r = ''.join(re.findall(r'[0-9]',cell.coordinate))
                            exact_r = int(exact_r) - 1
                            upper_cell_value = f'{exact_c}{exact_r}'
                            merged_ranges = sheet.merged_cells.ranges
                            for merged_range in merged_ranges:
                                top_left_cell = merged_range.min_row,merged_range.min_col
                                top_left_cell_address = f"{get_column_letter(top_left_cell[1])}{top_left_cell[0]}"
                                cell_in_range = list(merged_range.cells)
                                
                                for item in cell_in_range:
                                    r,c = item
                                    cell_address = f"{get_column_letter(c)}{r}"
                                    if upper_cell_value == cell_address:
                                        upper_cell_value = wb[sheet_name][top_left_cell_address].value
                            result_column.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'value': cell.value,
                                'criteria':str(upper_cell_value).lower().replace('\n',"")
                            })
        
        for col in sheet.iter_cols():
            for cell in col:
                for i, text in enumerate(search_queries):
                    search_text_row = search_queries[i]["search_query"]
                    if cell.value and search_text_row.lower() in str(cell.value).replace(" ","").lower().replace("\n",""):
                        result_row.append({
                            'sheet': sheet_name,
                            'cell': cell.coordinate,
                            'value': cell.value
                    })
                        num += 1
                        print(f"Found {num} out of {len(search_queries)} instruments in sheet")
    return result_row,result_column,search_queries
    

def query(row_query,column_query,file_path):
    wb = openpyxl.load_workbook(file_path)
    instrument_data = dict()
    for instrument in row_query:
        #data_dict = dict()
        some_list = []
        some_list = dict()
        sheet = instrument['sheet']
        name = instrument['value']
        cell_row = instrument['cell']
        #row_value =  instrument['value']
        exact_row = ''.join(re.findall(r'[0-9]',cell_row))
        for column in column_query:
            if sheet == column['sheet']:
                cell_column = column["cell"]
                value_column = str(column['value']).lower().replace("\n"," ").replace(" ","")
                exact_column = ''.join(re.findall(r'[A-Za-z]',cell_column))
                criteria_column = column['criteria'].replace("\n"," ").replace(" ","")

                wk_sheet = wb[sheet][f'{exact_column}{exact_row}'] 
                #some_list.append({f'{str(value_column)} {str(criteria_column)}':wk_sheet.value})
                some_list[f'{str(value_column)}-{str(criteria_column)}'] = wk_sheet.value

                column_index = column_index_from_string(exact_column) - 1
                column_ind = get_column_letter(column_index)
                alarm_uom = wb[sheet][f'{column_ind}{exact_row}']
                criteria = wb[sheet][f'{exact_column}{8}'].value
                if criteria != None and criteria.lower().replace(" ","") == 'alarm':
                    print(alarm_uom.value)
                    some_list["uom-alarm"] = alarm_uom.value
        instrument_data[f"{name}"] = some_list
    return instrument_data
    

def annnotate_pdf(source_doc,instrument_value,input_folder_path=None,output_folder_path=None):
    instrument_values = instrument_value.copy()
    to_docs = []

    instrument_classes = {
        "alarm":["high","low","high high","low low"], # ends with A
        "pressure":["normal operating"], #starts with P
        "temperature":["normal operating"], #starts with T
        "flowrate":["noraml operating"] #starts with F or Q
    }
    # the alarm criteria takes precedence
    for image in source_doc:
        for i,_ in enumerate(source_doc[image]):
            detections = source_doc[image][i][0]
            #bbox = source_doc[image][i][1]
            bbox_cord = source_doc[image][i][1]
            for instrument in instrument_values:
                if (detections !='') and (detections.lower() in instrument.lower().replace(" ","")):
                    instrument_values[instrument]["document"] = image
                    #instrument_values[instrument]["bbox_pixels"] = bbox
                    instrument_values[instrument]["bbox_cord"] = bbox_cord
                    break
    
    
    for instrument in instrument_values:
        main_doc = instrument_values[instrument]['document'].replace(".jpg","").split(".pdf_")
        x,y,w,h = instrument_values[instrument]['bbox_cord']
        nx,ny,nw,nh = normalized_to_pdf_rect_px_to_pt(x,y,w,h,int(2384),int(1684))
        pattern = r'[A-Z]{2,4}'
        matches = re.findall(pattern, instrument)

        while True:
            if matches[0].lower().endswith("a"):
                l_alarm = instrument_values[instrument]['low-alarm']
                ll_alarm = instrument_values[instrument]['lowlow-alarm']
                h_alarm = instrument_values[instrument]['high-alarm']
                hh_alarm = instrument_values[instrument]['highhigh-alarm']

                l = [l_alarm,ll_alarm,h_alarm,hh_alarm]
                uom = instrument_values[instrument]['uom-alarm']
                
                # open the document here and draw the important data
                #pdf_path = f"{input_folder_path}\{main_doc[0]}.pdf"
                #reader = PdfReader(pdf_path)
                #page = int(main_doc[1])
                text = ""
                for item in l:
                    if item != None:
                        text += f"{item} {uom}\n"
                #write_to_pdf(nx,ny,nw,nh,text,page,output_folder_path,reader)
                input_path = f"{input_folder_path}\{main_doc[0]}.pdf"
                output_path = f"{output_folder_path}\{main_doc[0]}.pdf"

                to_docs.append({"doc_name":main_doc[0],
                                "doc_path_in":input_path,
                                "doc_path_out":output_path,
                                "text":text,
                                "cord":(nx,ny,nw,nh),
                                "page":main_doc[1]})
                
                #annotate_doc(page,nx,ny,nw,nh,text)
                # we want to be able to do more than one annotation before the loop breaks
                

                # will have to add pdf opening to the same function as annotation
                # checks if the outfolder has a file with the same path as the output path
                # if it doesnt exist, it creates a new one and does the first annotation
                # in subsequent annotations it overwrites the file that was created 
                break

            elif matches[0].lower().startswith("p"):
                press = instrument_values[instrument]['normaloperating-pressureordifferentialpressure']
                uom = instrument_values[instrument]['uom-pressureordifferentialpressure']
                # open the document here and draw the important data
                input_path = f"{input_folder_path}\{main_doc[0]}.pdf"
                output_path = f"{output_folder_path}\{main_doc[0]}.pdf"
                text = ""
                text = f"{press} {uom}"
                to_docs.append({"doc_name":main_doc[0],
                                "doc_path_in":input_path,
                                "doc_path_out":output_path,
                                "text":text,
                             "cord":(nx,ny,nw,nh),
                             "page":main_doc[1]})
                break
                

            elif matches[0].lower().startswith("t"):
                temp = instrument_values[instrument]['normaloperating-temperature']
                uom = instrument_values[instrument]['uom-temperature']
                # open the document here and draw the important data
                input_path = f"{input_folder_path}\{main_doc[0]}.pdf"
                output_path = f"{output_folder_path}\{main_doc[0]}.pdf"
                text = ""
                text = f"{temp} {uom}"
                to_docs.append({"doc_name":main_doc[0],
                                "doc_path_in":input_path,
                                "doc_path_out":output_path,
                                "text":text,
                             "cord":(nx,ny,nw,nh),
                             "page":main_doc[1]})
                break
                
            elif matches[0].lower().startswith('f'):
                flow = instrument_values[instrument]['normal-flowrate']
                uom = instrument_values[instrument]['uom-flowrate']
                # open the document here and draw the important data
                input_path = f"{input_folder_path}\{main_doc[0]}.pdf"
                output_path = f"{output_folder_path}\{main_doc[0]}.pdf"
                text = ""
                text += f"{flow} {uom}"
                to_docs.append({"doc_name":main_doc[0],
                                "doc_path_in":input_path,
                                "doc_path_out":output_path,
                                "text":text,
                             "cord":(nx,ny,nw,nh),
                             "page":main_doc[1]})
                break
    
    return instrument_values,to_docs

def draw_on_doc(docs):
    for i,doc in enumerate(docs):
        print(f"Annotated {i+1} out of {len(docs)}")
        in_path = doc["doc_path_in"]
        out_path = doc["doc_path_out"]
        text = doc["text"]
        cord = doc["cord"]
        doc_page = int(doc['page'])
        y = 0 
        reader = PdfReader(in_path)
        writer = PdfWriter()

        for i, page in enumerate(reader.pages):
            if i == doc_page:
                page.rotate(90)

        for page in reader.pages:
            writer.add_page(page)

        # page = writer.pages[doc_page]
        # annotate_doc(page,cord[0],cord[1],cord[2],cord[3],text)

        for i, page in enumerate(writer.pages):
                pages = writer.pages[doc_page]
                if i == doc_page:
                    annotate_doc(pages,cord[0],cord[1],cord[2],cord[3],text)
                    page.rotate(-90)
                    

        #write_to_pdf(cord[0],cord[1],cord[2],cord[3],text,writer,doc_page)
        while y != 1:
            in_path, out_path = out_path, in_path
            with open(f"{out_path}", "wb") as f:
                writer.write(f)
            y += 1
            #in_path, out_path = out_path, in_path 
    
            
            
    
    


            


if __name__ == '__main__':
    pre_dict = detect_instrument(r"C:\Users\Angelo\Desktop\Holiday Stuff\Automation\test_run") 
    crop_dict = get_bbox(pre_dict)
    ocr_text = detect_text(crop_dict)
    formatted_text = format_text(ocr_text)
    row,column,search_queries = search_sheet(formatted_text,r"C:\Users\Angelo\Downloads\T-14.186.369-SH1-3_MK18.xlsx")
    instrument_data = query(row,column,r"C:\Users\Angelo\Downloads\T-14.186.369-SH1-3_MK18.xlsx")
    annotated,doc_list = annnotate_pdf(formatted_text,instrument_data,r"C:\Users\Angelo\Desktop\Holiday Stuff\Automation\Infolder",
                              r"C:\Users\Angelo\Desktop\Holiday Stuff\Automation\Outfolder")
    draw_on_doc(doc_list)
    #print(row)
    #print(column)
    #print(search_queries)
    #print(instrument_data["034PZA   -780 /P14"])
    #print(instrument_data)
    #print(annotated.keys())
    #print(annotated['034PZA   -780 /P14'])
    #print(doc_list)

    #img_plot = plt.imshow(formatted_text['Binder 1200-1300.pdf_5.jpg'][0][1])
    